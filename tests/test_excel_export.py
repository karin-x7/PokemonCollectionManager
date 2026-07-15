"""Tests for the Excel (.xlsx) export writer."""

from __future__ import annotations

from openpyxl import load_workbook

from app.export import excel_export
from app.export.models import COLUMNS, SEALED_COLUMNS, ExportRow, SealedExportRow

_ROW = ExportRow(
    collection_name="Binder",
    name="Xatu",
    set_name="Skyridge",
    card_number="H32",
    language="German",
    condition="Near Mint",
    extras="Reverse Holo",
    quantity=2,
    price=13.9,
    currency="EUR",
    price_quality="Exakter Treffer",
    price_updated_at="2026-07-04T12:00:00Z",
    notes="PSA 9",
    cardmarket_url="https://www.cardmarket.com/en/Pokemon/Products/Singles/Skyridge/Xatu",
)


def test_writes_header_and_row(tmp_path) -> None:
    path = tmp_path / "export.xlsx"

    excel_export.write([_ROW], path)

    workbook = load_workbook(path)
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    data_row = [cell.value for cell in sheet[2]]

    assert header == list(COLUMNS)
    assert data_row == [
        "Binder", "Xatu", "Skyridge", "H32", "German", "Near Mint", "Reverse Holo",
        2, "13,90", "EUR", "Exakter Treffer", "2026-07-04T12:00:00Z", "PSA 9",
        "https://www.cardmarket.com/en/Pokemon/Products/Singles/Skyridge/Xatu",
    ]


def test_header_row_is_bold(tmp_path) -> None:
    path = tmp_path / "export.xlsx"

    excel_export.write([_ROW], path)

    sheet = load_workbook(path).active
    assert all(cell.font.bold for cell in sheet[1])


def test_empty_rows_still_writes_header(tmp_path) -> None:
    path = tmp_path / "export.xlsx"

    excel_export.write([], path)

    sheet = load_workbook(path).active
    assert [cell.value for cell in sheet[1]] == list(COLUMNS)
    assert sheet.max_row == 1


def test_write_sealed_writes_header_and_row(tmp_path) -> None:
    path = tmp_path / "export.xlsx"
    row = SealedExportRow(
        name="Base Set Booster Box",
        category="Booster Box",
        language="German",
        quantity=1,
        price=5000.0,
        currency="EUR",
        price_quality="Exakter Treffer",
        price_updated_at="2026-07-05T00:00:00Z",
        notes="",
        cardmarket_url="https://www.cardmarket.com/en/Pokemon/Products/Booster-Boxes/Base-Set-Booster-Box",
    )

    excel_export.write_sealed([row], path)

    workbook = load_workbook(path)
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    data_row = [cell.value for cell in sheet[2]]

    assert header == list(SEALED_COLUMNS)
    assert data_row == [
        # openpyxl round-trips a written "" as None, not "" -- same as an
        # empty "Notizen" cell would for a card export.
        "Base Set Booster Box", "Booster Box", "German", 1,
        "5.000,00", "EUR", "Exakter Treffer", "2026-07-05T00:00:00Z", None,
        "https://www.cardmarket.com/en/Pokemon/Products/Booster-Boxes/Base-Set-Booster-Box",
    ]

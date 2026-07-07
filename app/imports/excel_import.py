"""Excel (.xlsx) import -- reads the same layout ``app.export.excel_export`` writes."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.imports.models import ImportedCardRow, ImportedSealedRow, ImportFileError


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read(path: Path) -> list[ImportedCardRow]:
    """Read owned-card rows from the first sheet of an Excel workbook.

    Only the columns actually needed are looked up by header name (not
    position), so a file with extra/reordered/missing columns still works.
    """
    rows = _read_dicts(path)
    return [
        ImportedCardRow(
            collection_name=row.get("Sammlung", ""),
            name=row.get("Name", ""),
            set_name=row.get("Set", ""),
            card_number=row.get("Nr.", ""),
            language=row.get("Sprache", ""),
            condition=row.get("Zustand", ""),
            extras=row.get("Extra", ""),
            quantity=row.get("Menge", ""),
            notes=row.get("Notizen", ""),
            cardmarket_url=row.get("Cardmarket-Link", ""),
        )
        for row in rows
    ]


def read_sealed(path: Path) -> list[ImportedSealedRow]:
    """Read owned-sealed-product rows from the first sheet of an Excel workbook."""
    rows = _read_dicts(path)
    return [
        ImportedSealedRow(
            name=row.get("Name", ""),
            category=row.get("Kategorie", ""),
            language=row.get("Sprache", ""),
            quantity=row.get("Menge", ""),
            notes=row.get("Notizen", ""),
            cardmarket_url=row.get("Cardmarket-Link", ""),
        )
        for row in rows
    ]


def _read_dicts(path: Path) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, OSError) as exc:
        raise ImportFileError(f"Could not read '{path.name}': {exc}") from exc

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [_cell_text(cell) for cell in next(rows_iter)]
    except StopIteration:
        return []
    return [
        dict(zip(header, (_cell_text(cell) for cell in row)))
        for row in rows_iter
        if any(cell is not None for cell in row)
    ]

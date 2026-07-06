"""Excel (.xlsx) export."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.export.models import COLUMNS, SEALED_COLUMNS, ExportRow, SealedExportRow

_SHEET_TITLE = "Sammlung"


def write(rows: list[ExportRow], path: Path) -> None:
    """Write ``rows`` to ``path`` as a single-sheet Excel workbook with a

    bold header row and columns auto-sized to their content."""
    _write(rows, COLUMNS, path)


def write_sealed(rows: list[SealedExportRow], path: Path) -> None:
    """Write sealed-product ``rows`` to ``path`` as a single-sheet Excel workbook."""
    _write(rows, SEALED_COLUMNS, path)


def _write(
    rows: list[ExportRow] | list[SealedExportRow], columns: tuple[str, ...], path: Path
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET_TITLE

    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(row.as_tuple())

    for col_index, header in enumerate(columns, start=1):
        widths = [len(header)] + [
            len(str(row.as_tuple()[col_index - 1])) for row in rows
        ]
        sheet.column_dimensions[get_column_letter(col_index)].width = min(max(widths) + 2, 60)

    workbook.save(path)
